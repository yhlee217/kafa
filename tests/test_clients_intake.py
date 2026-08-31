"""수임처 속성 조사표 — 양식 생성 / 파싱 / YAML 변환."""
import openpyxl
import pytest

from kafa.clients import COLUMNS, SHEET, parse_template, to_yaml, write_template


def test_template_has_sheets_and_header(tmp_path):
    p = write_template(tmp_path / "t.xlsx", ["고객001", "고객002"])
    wb = openpyxl.load_workbook(p)
    assert SHEET in wb.sheetnames and "작성방법" in wb.sheetnames
    ws = wb[SHEET]
    assert [c.value for c in ws[1]] == COLUMNS
    assert [ws.cell(row=r, column=1).value for r in (2, 3)] == ["고객001", "고객002"]


def _filled(tmp_path, rows):
    p = write_template(tmp_path / "f.xlsx")
    wb = openpyxl.load_workbook(p)
    ws = wb[SHEET]
    for r in rows:
        ws.append(r)
    wb.save(p)
    return p


def test_parse_maps_korean_answers(tmp_path):
    p = _filled(tmp_path, [
        ["1인사업자", "개인", "아니오", "카페", "혼자 운영"],
        ["행복상사", "법인", "예", "", ""],
    ])
    got = parse_template(p)
    assert got["1인사업자"]["client_type"] == "individual"
    assert got["1인사업자"]["has_employees"] is False
    assert "카페" in got["1인사업자"]["note"]
    assert got["행복상사"] == {"client_type": "corporate", "has_employees": True}


def test_parse_tolerates_variants_and_blanks(tmp_path):
    p = _filled(tmp_path, [
        ["A", "법인사업자", "있음", "", ""],
        ["B", "", "", "", ""],            # 미기재 → 빈 속성(기본값 사용)
        ["", "개인", "예", "", ""],        # 이름 없음 → 무시
    ])
    got = parse_template(p)
    assert got["A"]["has_employees"] is True and got["A"]["client_type"] == "corporate"
    assert got["B"] == {}
    assert "" not in got


def test_to_yaml_roundtrips(tmp_path):
    import yaml
    text = to_yaml({"고객001": {"client_type": "individual", "has_employees": False}})
    data = yaml.safe_load(text)
    assert data["defaults"]["client_type"] == "corporate"
    assert data["clients"]["고객001"]["has_employees"] is False
    # config_loader 가 그대로 읽을 수 있어야 한다
    d = tmp_path / "config"; d.mkdir()
    (d / "clients.yaml").write_text(text, encoding="utf-8")
    from kafa.config_loader import client_profile, load_clients
    load_clients.cache_clear()
    prof = client_profile("고객001", str(d))
    assert prof["client_type"] == "individual" and prof["has_employees"] is False
    load_clients.cache_clear()


# ── 이름 미리 채우기 (담당자가 타이핑하지 않게) ──

def test_names_from_excel_finds_column_by_header(tmp_path):
    """헤더 위치·컬럼 순서를 몰라도 이름 컬럼을 구조로 찾는다."""
    p = tmp_path / "거래처목록.xlsx"
    wb = openpyxl.Workbook(); ws = wb.active
    ws.append(["거래처 목록 (2026년)"])          # 제목행
    ws.append([])
    ws.append(["코드", "거래처명", "사업자번호"])  # 실제 헤더
    ws.append(["001", "행복상사", "111-11-11119"])
    ws.append(["002", "대박유통", "222-22-22228"])
    ws.append(["", "합계", ""])                   # 집계행 → 제외
    ws.append(["003", "행복상사", ""])            # 중복 → 제외
    wb.save(p)
    from kafa.clients import names_from_excel
    assert names_from_excel(p) == ["행복상사", "대박유통"]


def test_names_from_excel_ignores_sheet_without_name_column(tmp_path):
    p = tmp_path / "x.xlsx"
    wb = openpyxl.Workbook(); ws = wb.active
    ws.append(["금액", "일자"]); ws.append([1000, "01-01"])
    wb.save(p)
    from kafa.clients import names_from_excel
    assert names_from_excel(p) == []


def test_names_from_inbox_uses_folder_names(tmp_path):
    inbox = tmp_path / "inbox"
    for n in ("고객002", "고객001", "_archive"):
        (inbox / n).mkdir(parents=True)
    from kafa.clients import names_from_inbox
    assert names_from_inbox(inbox) == ["고객001", "고객002"]   # _로 시작하면 제외


def test_template_prefills_names(tmp_path):
    p = write_template(tmp_path / "t.xlsx", ["행복상사", "대박유통"])
    ws = openpyxl.load_workbook(p)[SHEET]
    assert [ws.cell(row=r, column=1).value for r in (2, 3)] == ["행복상사", "대박유통"]
    # 이름만 채우고 나머지는 비워 둔다(담당자가 고르도록)
    assert ws.cell(row=2, column=2).value in (None, "")


def test_picks_best_sheet_not_all(tmp_path):
    """요약·점검 같은 보조 시트를 함께 긁어 이름이 뒤섞이면 안 된다."""
    p = tmp_path / "master.xlsx"
    wb = openpyxl.Workbook()
    s0 = wb.active; s0.title = "00_요약"
    s0.append(["수임처 마스터"]); s0.append([])
    s0.append(["항목", "값"]); s0.append(["총수임처", 3]); s0.append(["법인", 1])
    s1 = wb.create_sheet("10_마스터")
    s1.append(["No", "회사명", "구분"])
    for i, (n, t) in enumerate([("가상상사", "법인"), ("나린유통", "개인"),
                                ("다온물산", "개인")], 1):
        s1.append([i, n, t])
    s2 = wb.create_sheet("20_점검필요")
    s2.append(["회사명", "유형"]); s2.append(["나린유통", "사업자번호 없음"])
    wb.save(p)

    from kafa.clients import profiles_from_excel
    got = profiles_from_excel(p)
    assert [g["name"] for g in got] == ["가상상사", "나린유통", "다온물산"]
    assert got[0]["client_type"] == "법인" and got[1]["client_type"] == "개인"


def test_template_prefills_client_type(tmp_path):
    p = write_template(tmp_path / "t.xlsx",
                       [{"name": "가상상사", "client_type": "법인"}, {"name": "나린유통"}])
    ws = openpyxl.load_workbook(p)[SHEET]
    assert [ws.cell(row=2, column=c).value for c in (1, 2, 3)] == ["가상상사", "법인", None]
    assert ws.cell(row=3, column=2).value in (None, "")   # 모르면 비워 둔다


# ── 수임처 마스터 엑셀에서 수임처코드 뽑기 ──

def test_client_codes_extracted_from_master_urls(tmp_path):
    """접속 URL 안의 cno 가 목록에서 수임처를 여는 데 쓰는 코드다."""
    import openpyxl

    from kafa.clients import client_cnos, client_cnos_from_excel

    path = tmp_path / "master.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["회사명", "구분", "접속 URL"])
    ws.append(["행복상사", "법인",
               "https://smarta.wehago.com/#/smarta/account/SAAC0105"
               "?sao&cno=10049328&cd_com=biz1&gisu=15"])
    ws.append(["코드없는곳", "법인", "https://smarta.wehago.com/#/main"])
    wb.save(path)

    assert client_cnos_from_excel(path) == {"행복상사": "10049328"}
    assert client_cnos(path) == {"행복상사": "10049328"}


# ── 수임처 마스터 → clients.yaml 자동 채우기 ──

def _master(tmp_path, rows):
    import openpyxl
    path = tmp_path / "master.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["회사명", "구분", "접속 URL"])
    for name, gubun in rows:
        ws.append([name, gubun, "https://x/?cno=1"])
    wb.save(path)
    return path


def test_master_fills_client_type(tmp_path):
    from kafa.clients import profiles_from_master
    path = _master(tmp_path, [("행복상사", "법인"), ("김아무개", "개인")])
    got = profiles_from_master(path)
    assert got["행복상사"]["client_type"] == "corporate"
    assert got["김아무개"]["client_type"] == "individual"
    # 직원 유무는 자료로 알 수 없다 — 넣지 않는다
    assert "has_employees" not in got["행복상사"]


def test_master_key_matches_inbox_folder_name(tmp_path):
    """client_id 는 파이프라인 고객 폴더명(safe_name)과 같아야 한다."""
    from kafa.clients import profiles_from_master
    from kafa.fetch.plan import safe_name
    path = _master(tmp_path, [("가/나 상사", "법인")])
    key = next(iter(profiles_from_master(path)))
    assert key == safe_name("가/나 상사") and "/" not in key


def test_merge_keeps_human_answers(tmp_path):
    """다시 돌려도 담당자가 적은 직원 유무·비고를 지우지 않는다."""
    from kafa.clients import merge_profiles
    existing = {"행복상사": {"client_type": "corporate", "has_employees": False,
                          "note": "1인 사업자"}}
    merged = merge_profiles(existing, {"행복상사": {"client_type": "individual",
                                                "name": "행복상사"}})
    assert merged["행복상사"]["has_employees"] is False
    assert merged["행복상사"]["note"] == "1인 사업자"
    assert merged["행복상사"]["client_type"] == "individual"   # 마스터 값은 갱신


def test_cli_from_master_writes_yaml(tmp_path):
    from kafa.clients_cli import main
    path = _master(tmp_path, [("행복상사", "법인"), ("김아무개", "개인")])
    out = tmp_path / "clients.yaml"
    assert main(["from-master", str(path), "--out", str(out)]) == 0
    import yaml
    data = yaml.safe_load(out.read_text(encoding="utf-8"))
    assert data["clients"]["김아무개"]["client_type"] == "individual"
    assert data["defaults"]["client_type"] == "corporate"


def test_cli_from_master_is_rerunnable(tmp_path):
    """두 번 돌려도 사람이 채운 값이 살아 있다."""
    from kafa.clients_cli import main
    path = _master(tmp_path, [("행복상사", "법인")])
    out = tmp_path / "clients.yaml"
    main(["from-master", str(path), "--out", str(out)])
    import yaml
    data = yaml.safe_load(out.read_text(encoding="utf-8"))
    data["clients"]["행복상사"]["has_employees"] = False
    out.write_text(yaml.safe_dump(data, allow_unicode=True), encoding="utf-8")
    main(["from-master", str(path), "--out", str(out)])
    again = yaml.safe_load(out.read_text(encoding="utf-8"))
    assert again["clients"]["행복상사"]["has_employees"] is False


def test_cli_from_master_missing_file(tmp_path):
    from kafa.clients_cli import main
    assert main(["from-master", str(tmp_path / "없음.xlsx")]) == 2
