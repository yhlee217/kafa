"""계정과목(참고용) 시트 파서 — 구조(헤더) 기반. 합성 시트만 사용."""
import openpyxl
import pytest

from kafa.config_loader import load_account_codes
from kafa.io_wehago.account_sheet import build_mapping, parse_account_sheet


def _sheet(path, header, rows, sheet_title="계정과목(참고용)"):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = sheet_title
    ws.append(header)
    for r in rows:
        ws.append(r)
    wb.save(path)
    return path


def test_parse_basic(tmp_path):
    p = _sheet(tmp_path / "form.xlsx", ["계정과목", "계정코드"],
               [["미지급비용", 262], ["(판)복리후생비", 811], ["(판)차량유지비", 822]])
    m = parse_account_sheet(p)
    assert m == {"미지급비용": 262, "(판)복리후생비": 811, "(판)차량유지비": 822}


def test_header_variants_and_skip_nonnumeric(tmp_path):
    # 헤더가 '계정명'/'코드' 이고, 코드가 숫자 아닌 머리글/소계 행은 건너뛴다
    p = _sheet(tmp_path / "form2.xlsx", ["계정명", "코드"],
               [["구분", "코드"],            # 머리글 반복 → 코드 비숫자 → 제외
                ["현금", "101"],
                ["소계", ""],                # 코드 없음 → 제외
                ["외상매입금", 251]])
    m = parse_account_sheet(p)
    assert m == {"현금": 101, "외상매입금": 251}


def test_picks_named_sheet_among_many(tmp_path):
    wb = openpyxl.Workbook()
    wb.active.title = "표지"
    wb.active.append(["아무거나", "값"])
    ws2 = wb.create_sheet("계정과목")
    ws2.append(["계정과목", "계정코드"])
    ws2.append(["선급금", 131])
    p = tmp_path / "multi.xlsx"
    wb.save(p)
    assert parse_account_sheet(p) == {"선급금": 131}


def test_code_column_priority_over_decoy(tmp_path):
    # '관리코드' 디코이가 있어도 '계정코드'를 우선 선택해야 한다
    p = _sheet(tmp_path / "form_decoy.xlsx", ["계정과목", "관리코드", "계정코드"],
               [["현금", "A1", 101], ["미지급비용", "B2", 262]])
    assert parse_account_sheet(p) == {"현금": 101, "미지급비용": 262}


def test_missing_columns_raises(tmp_path):
    p = _sheet(tmp_path / "bad.xlsx", ["이름", "값"], [["x", 1]])
    with pytest.raises(ValueError):
        parse_account_sheet(p)


def test_build_mapping_merges_config_first(tmp_path):
    # 시트가 같은 계정에 다른 코드를 줘도 config(검증분)가 우선
    p = _sheet(tmp_path / "form3.xlsx", ["계정과목", "계정코드"],
               [["미지급비용", 999], ["임차료", 819]])
    m = build_mapping(p)
    assert m["미지급비용"] == 262          # config 우선(999 무시)
    assert m["임차료"] == 819              # 시트에서 보강된 신규 계정


def test_build_mapping_bad_sheet_is_safe(tmp_path):
    p = _sheet(tmp_path / "bad2.xlsx", ["이름", "값"], [["x", 1]])
    m = build_mapping(p)                    # 컬럼 못 찾아도 예외 없이 config만
    assert m["미지급비용"] == 262


def _pair_sheet(path, sheet_title="계정과목(참고용)"):
    """실제 양식 구조: 분류행(자산/제조/판관비...) + (코드,계정과목) 쌍이 가로 반복."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = sheet_title
    ws.append(["자산", None, "제조", None, "판관비", None])
    ws.append(["코드", "계정과목", "코드", "계정과목", "코드", "계정과목"])
    ws.append([101, "현금", 511, "복리후생비", 811, "복리후생비"])
    ws.append([102, "당좌예금", 512, "여비교통비", 812, "여비교통비"])
    wb.save(path)
    return path


def test_parse_pair_repeating_header_applies_category_markers(tmp_path):
    p = _pair_sheet(tmp_path / "pairs.xlsx")
    m = parse_account_sheet(p)
    # 원가구분(제조/도급/분양/판관비)만 접두 마커, 자산/부채/매출/영업외는 마커 없음
    assert m == {
        "현금": 101, "당좌예금": 102,
        "(제)복리후생비": 511, "(제)여비교통비": 512,
        "(판)복리후생비": 811, "(판)여비교통비": 812,
    }


def test_pair_header_takes_priority_over_simple_fallback(tmp_path):
    # 쌍 반복 구조가 있으면(마커 적용) 단순 컬럼명 탐지로 빠지지 않는다
    p = _pair_sheet(tmp_path / "pairs2.xlsx")
    m = parse_account_sheet(p)
    assert "(제)복리후생비" in m
    assert "복리후생비" not in m          # 마커 없는 단순 매핑으로 오인되지 않음


def test_pair_header_empty_data_falls_back_and_raises(tmp_path):
    # 쌍 헤더는 찾았지만 데이터 행이 전부 공란이면(코드 없음) 단순 구조로 폴백하되,
    # 분류행은 단순 헤더로도 못 쓰이므로 결국 컬럼을 못 찾아 예외.
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "계정과목(참고용)"
    ws.append(["자산", None])
    ws.append(["코드", "계정과목"])
    wb.save(tmp_path / "pairs_empty.xlsx")
    with pytest.raises(ValueError):
        parse_account_sheet(tmp_path / "pairs_empty.xlsx")


def test_no_matching_sheet_name_falls_back_to_first_sheet(tmp_path):
    p = _sheet(tmp_path / "unnamed.xlsx", ["계정과목", "계정코드"],
               [["현금", 101]], sheet_title="Sheet1")
    assert parse_account_sheet(p) == {"현금": 101}


def test_build_mapping_merges_pair_structure_sheet(tmp_path):
    p = _pair_sheet(tmp_path / "pairs3.xlsx")
    m = build_mapping(p)
    assert m["미지급비용"] == 262           # config 검증분 그대로 보존
    assert m["(제)복리후생비"] == 511       # 시트에서 보강


def _config_dir(tmp_path, sheet_path):
    """account_sheet_path 가 연결된 임시 config 디렉터리 생성."""
    d = tmp_path / "config"
    d.mkdir()
    (d / "account_codes.yaml").write_text(
        "account_name_to_code:\n  미지급비용: 262\n", encoding="utf-8")
    (d / "rules.yaml").write_text(
        f'account_sheet_path: "{sheet_path}"\n', encoding="utf-8")
    return str(d)


def test_load_account_codes_merges_sheet_via_config(tmp_path):
    sheet = _sheet(tmp_path / "form_cfg.xlsx", ["계정과목", "계정코드"],
                   [["미지급비용", 999], ["임차료", 819]])
    m = load_account_codes(_config_dir(tmp_path, sheet))
    assert m["미지급비용"] == 262          # config 검증분 우선(999 무시)
    assert m["임차료"] == 819              # 시트에서 자동 보강


def test_load_account_codes_no_sheet(tmp_path):
    d = tmp_path / "config2"
    d.mkdir()
    (d / "account_codes.yaml").write_text(
        "account_name_to_code:\n  현금: 101\n", encoding="utf-8")
    (d / "rules.yaml").write_text("account_sheet_path: null\n", encoding="utf-8")
    m = load_account_codes(str(d))
    assert m == {"현금": 101}              # 시트 미지정 → config 그대로
