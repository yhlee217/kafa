"""입력 형식 견고성 — 위하고 파일 아닌 경우 명확한 에러."""
import openpyxl
import pytest

from kafa.io_wehago.reader import InputFormatError, read_download_xlsx


def _xlsx(tmp_path, header, rows=()):
    wb = openpyxl.Workbook(); ws = wb.active
    ws.append(header)
    for r in rows:
        ws.append(r)
    p = tmp_path / "f.xlsx"
    wb.save(p)
    return p


def test_missing_required_columns_raises(tmp_path):
    p = _xlsx(tmp_path, ["엉뚱", "컬럼"])
    with pytest.raises(InputFormatError) as e:
        read_download_xlsx(p)
    assert "필수 컬럼" in str(e.value)


def test_valid_header_ok(tmp_path):
    header = ["연도", "일자", "Code", "거래처", "구분", "품명", "공급가액", "세액",
              "비과세", "합계", "국세청", "업태", "종목", "유형", "차변계정",
              "대변계정", "관리", "전표상태", "사업자등록번호"]
    p = _xlsx(tmp_path, header,
              [["2026", "03-15", "A", "가맹점", "법인", "커피", 5000, 500, 0, 5500,
                "공제", "음식점", "카페", "카과", "(판)복리후생비", "", "", "", "111-11-11119"]])
    rows = read_download_xlsx(p)
    assert len(rows) == 1
    assert rows[0].거래처 == "가맹점"


# ── 실제 파일(2026-08 샘플)에서 발견된 회귀 ──

def test_summary_rows_in_date_column_are_excluded(tmp_path):
    """요약 문구가 '일자' 칸에 오는 실제 형식을 걸러낸다(거래처는 비어 있음)."""
    import openpyxl
    from kafa.io_wehago.schema import INPUT_COLUMNS
    p = tmp_path / "sum.xlsx"
    wb = openpyxl.Workbook(); ws = wb.active
    ws.append(INPUT_COLUMNS)
    row = ["2026", "01-02", "A", "가맹점", "법인", "품명", 1000, 100, 0, 1100,
           "공제", "도매", "문구", "카과", "(판)소모품비", "미지급비용", "", "확정가능",
           "111-11-11119"]
    ws.append(row)
    for label, total in [("카드사별  매입 : 11건", "341,900"),
                         ("카드사별  일반 : 65건", "5,554,945"),
                         ("합계(카드사:2) : 82건", "6057533")]:
        s = [""] * len(INPUT_COLUMNS)
        s[1] = label; s[9] = total
        ws.append(s)
    wb.save(p)
    rows = read_download_xlsx(p)
    assert len(rows) == 1                    # 요약 3행 제외


def test_pending_account_literal_becomes_blank(tmp_path):
    """차변계정 칸의 '미추천' 문구는 계정명이 아니므로 공란으로 정규화."""
    import openpyxl
    from kafa.io_wehago.schema import INPUT_COLUMNS
    p = tmp_path / "pend.xlsx"
    wb = openpyxl.Workbook(); ws = wb.active
    ws.append(INPUT_COLUMNS)
    ws.append(["2026", "01-02", "A", "가맹점", "법인", "", 1000, 100, 0, 1100,
               "불공제", "음식점업", "한식", "일반", "미추천", "미지급비용", "",
               "미추천", "111-11-11119"])
    wb.save(p)
    r = read_download_xlsx(p)[0]
    assert r.차변계정 == "" and r.전표상태 == "미추천"


def test_blank_amount_cell_is_zero_not_nan(tmp_path):
    """빈 금액 셀이 Decimal('NaN')이 되면 이후 모든 비교가 터진다 → 0 으로."""
    import openpyxl
    from decimal import Decimal
    from kafa.io_wehago.schema import INPUT_COLUMNS
    p = tmp_path / "nan.xlsx"
    wb = openpyxl.Workbook(); ws = wb.active
    ws.append(INPUT_COLUMNS)
    ws.append(["2026", "01-02", "A", "가맹점", "법인", "품명", 1000, 100, None, 1100,
               "공제", "도매", "문구", "카과", "(판)소모품비", "미지급비용", "",
               "확정가능", "111-11-11119"])
    wb.save(p)
    r = read_download_xlsx(p)[0]
    assert r.비과세 == Decimal(0) and r.비과세.is_finite()
    assert (r.비과세 > 0) is False           # 비교가 예외 없이 동작
