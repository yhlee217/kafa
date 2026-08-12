"""수임처 속성 조사표 — 엑셀로 받아 config/clients.yaml 로 바꾼다.

담당자만 아는 사실(개인/법인·직원 유무)을 자료로는 알 수 없어 사람이 적어야 한다.
YAML 을 직접 편집하게 하는 대신, **엑셀 양식**을 채워 오면 변환한다(실무자 친화).

배경·근거: docs/domain_notes.md ("직원이 없으면 복리후생비가 성립하지 않는다" 등).
"""
from __future__ import annotations

from pathlib import Path

SHEET = "수임처"
COLUMNS = ["수임처 이름", "개인/법인", "직원 있나요?", "업종(선택)", "비고(선택)"]

_YES = {"예", "있음", "y", "yes", "o", "1", "true"}
_NO = {"아니오", "아니요", "없음", "n", "no", "x", "0", "false"}
_INDIVIDUAL = {"개인", "개인사업자", "individual"}
_CORPORATE = {"법인", "법인사업자", "corporate"}

_HOWTO = [
    ["수임처 속성 조사표 — 작성 방법"],
    [],
    ["이 표는 프로그램이 전표를 더 정확히 분류하기 위해 필요합니다."],
    ["자료(엑셀)만으로는 알 수 없고, 담당자만 아는 사실이라 직접 적어 주셔야 합니다."],
    [],
    ["1. '수임처' 시트에 거래처를 한 줄에 하나씩 적어 주세요."],
    ["2. '개인/법인' 과 '직원 있나요?' 는 칸을 클릭하면 목록에서 고를 수 있습니다."],
    ["3. 업종·비고는 안 적으셔도 됩니다."],
    [],
    ["왜 '직원 있나요?' 를 묻나요"],
    ["  직원이 없으면 복리후생비가 성립하지 않아, 마트·편의점·식사 같은 소액이"],
    ["  접대비로 가야 합니다. 이걸 알면 프로그램이 계정을 훨씬 잘 맞춥니다."],
    [],
    ["왜 '개인/법인' 을 묻나요"],
    ["  법인은 카드 부채(미지급비용)를 반드시 잡아야 하고, 개인은 인출금으로 갑니다."],
    ["  처리 방식이 아예 달라서 구분이 필요합니다."],
    [],
    ["※ 사업자번호·거래처 실명 같은 민감정보는 적지 않으셔도 됩니다."],
]


def write_template(path, clients=None) -> Path:
    """조사표(.xlsx) 생성.

    clients: 이름 문자열 목록, 또는 {"name":…, "client_type": "개인"|"법인"} 목록.
    알고 있는 값은 미리 채워 담당자가 고를 칸을 최소화한다.
    """
    import openpyxl
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.worksheet.datavalidation import DataValidation

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = SHEET
    ws.append(COLUMNS)
    head = PatternFill("solid", fgColor="E4F0EE")
    for c in range(1, len(COLUMNS) + 1):
        cell = ws.cell(row=1, column=c)
        cell.font = Font(bold=True)
        cell.fill = head
        cell.alignment = Alignment(horizontal="center")
    for c in (clients or []):
        if isinstance(c, dict):
            ws.append([c.get("name", ""), c.get("client_type", ""), "", "", ""])
        else:
            ws.append([c, "", "", "", ""])

    rows = max(len(clients or []) + 1, 200)
    dv_type = DataValidation(type="list", formula1='"개인,법인"', allow_blank=True)
    dv_emp = DataValidation(type="list", formula1='"예,아니오"', allow_blank=True)
    ws.add_data_validation(dv_type)
    ws.add_data_validation(dv_emp)
    dv_type.add(f"B2:B{rows}")
    dv_emp.add(f"C2:C{rows}")

    for col, width in zip("ABCDE", (24, 12, 14, 18, 34)):
        ws.column_dimensions[col].width = width
    ws.freeze_panes = "A2"

    how = wb.create_sheet("작성방법")
    for line in _HOWTO:
        how.append(line)
    how.column_dimensions["A"].width = 78
    how["A1"].font = Font(bold=True, size=13)

    out = Path(path)
    out.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out)
    return out


def _norm(v) -> str:
    return "" if v is None else str(v).strip()


# 이름 컬럼 탐지 키워드(구체적인 것 우선). 값을 추측하지 않고 **구조**로 찾는다.
NAME_HEADER_KEYS = ("거래처명", "수임처명", "사업장명", "업체명", "회사명",
                    "상호", "거래처", "수임처", "성명", "이름")
_SKIP_VALUES = {"합계", "소계", "총계", "계"}
TYPE_HEADER_KEYS = ("구분", "개인법인", "법인구분", "사업자구분", "유형")
URL_HEADER_KEYS = ("접속URL", "URL", "링크", "주소", "접속주소")


def _scan_rows(ws, limit: int = 5000) -> list:
    rows = []
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        rows.append(row)
        if i > limit:
            break
    return rows


def _header_in(rows, max_header_scan: int):
    """(점수, 이름컬럼, 구분컬럼, 헤더행) — 못 찾으면 점수 0.

    긴 문장(제목·안내문)은 헤더로 보지 않고, '회사명' 같은 정확 일치를 부분 일치보다
    우선한다. 이렇게 해야 '총수임처' 같은 요약 문구에 낚이지 않는다.
    """
    for i, row in enumerate(rows[:max_header_scan]):
        best_score, best_col, tcol, ucol = 0, None, None, None
        for j, cell in enumerate(row or []):
            text = _norm(cell).replace(" ", "")
            if not text or len(text) > 10:
                continue
            if text in NAME_HEADER_KEYS:
                score = 3
            elif any(k in text for k in NAME_HEADER_KEYS):
                score = 2
            else:
                if text in TYPE_HEADER_KEYS and tcol is None:
                    tcol = j
                elif text.upper() in URL_HEADER_KEYS and ucol is None:
                    ucol = j
                continue
            if score > best_score:
                best_score, best_col = score, j
        if best_col is not None:
            return best_score, best_col, tcol, ucol, i
    return 0, -1, None, None, 0


def _best_table(path, *, max_header_scan: int = 20):
    """여러 시트 중 **가장 그럴듯한 표 하나**를 고른다.

    정확 일치 헤더(점수 3)를 우선하고, 같은 점수면 데이터 행이 많은 시트를 쓴다.
    요약·점검 같은 보조 시트를 함께 긁어 이름이 뒤섞이는 것을 막는다.
    """
    import openpyxl

    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    best = None
    for ws in wb.worksheets:
        rows = _scan_rows(ws)
        score, ncol, tcol, ucol, hi = _header_in(rows, max_header_scan)
        if not score:
            continue
        n = sum(1 for r in rows[hi + 1:]
                if r and len(r) > ncol and _norm(r[ncol])
                and _norm(r[ncol]) not in _SKIP_VALUES)
        key = (score, n)
        if best is None or key > best[0]:
            best = (key, rows, ncol, tcol, ucol, hi)
    return best


def profiles_from_excel(path, *, max_header_scan: int = 20) -> list[dict]:
    """수임처 목록 엑셀 → [{"name":…, "client_type":…}].

    이름 컬럼은 헤더를 보고 **구조로** 찾고, 같은 표에 '구분' 컬럼이 있으면 개인/법인도
    함께 읽는다(담당자가 고를 칸이 줄어든다). 중복·합계행은 제외하고 순서를 유지한다.
    """
    best = _best_table(path, max_header_scan=max_header_scan)
    if best is None:
        return []
    _key, rows, ncol, tcol, ucol, hi = best
    out: list[dict] = []
    seen: set[str] = set()
    for row in rows[hi + 1:]:
        name = _norm(row[ncol] if row and len(row) > ncol else "")
        if not name or name in _SKIP_VALUES or name in seen:
            continue
        seen.add(name)
        rec = {"name": name}
        if tcol is not None and row and len(row) > tcol:
            t = _norm(row[tcol])
            if t in _INDIVIDUAL:
                rec["client_type"] = "개인"
            elif t in _CORPORATE:
                rec["client_type"] = "법인"
        if ucol is not None and row and len(row) > ucol:
            url = _norm(row[ucol])
            if url.startswith("http"):
                rec["url"] = url
        out.append(rec)
    return out


def client_urls_from_excel(path) -> dict[str, str]:
    """수임처 마스터 → {회사명: 접속 URL}. URL 컬럼이 없으면 빈 dict.

    로그인은 사람이 하고, **이미 로그인된 세션**에서 이 URL로 바로 이동해 화면 탐색
    단계를 줄인다(거래처 검색·선택 클릭 불필요).
    """
    return {r["name"]: r["url"] for r in profiles_from_excel(path) if r.get("url")}


def names_from_excel(path, *, max_header_scan: int = 20) -> list[str]:
    """거래처/수임처 목록 엑셀에서 이름만 뽑는다."""
    return [r["name"] for r in
            profiles_from_excel(path, max_header_scan=max_header_scan)]


def names_from_inbox(inbox) -> list[str]:
    """inbox 하위 폴더명 = 고객 ID(파이프라인 규칙)."""
    p = Path(inbox)
    if not p.exists():
        return []
    return sorted(d.name for d in p.iterdir()
                  if d.is_dir() and not d.name.startswith("_"))


def names_from_text(path) -> list[str]:
    """한 줄에 하나씩 적은 목록 파일."""
    lines = Path(path).read_text(encoding="utf-8").splitlines()
    return [ln.strip() for ln in lines if ln.strip() and not ln.startswith("#")]


def parse_template(path) -> dict[str, dict]:
    """채워진 조사표 → {수임처: {client_type, has_employees, note}}."""
    import openpyxl

    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb[SHEET] if SHEET in wb.sheetnames else wb.worksheets[0]
    out: dict[str, dict] = {}
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i == 0:
            continue                       # 헤더
        name = _norm(row[0] if len(row) > 0 else "")
        if not name:
            continue
        ctype = _norm(row[1] if len(row) > 1 else "").lower()
        emp = _norm(row[2] if len(row) > 2 else "").lower()
        업종 = _norm(row[3] if len(row) > 3 else "")
        비고 = _norm(row[4] if len(row) > 4 else "")

        prof: dict = {}
        if ctype in _INDIVIDUAL:
            prof["client_type"] = "individual"
        elif ctype in _CORPORATE:
            prof["client_type"] = "corporate"
        if emp in _YES:
            prof["has_employees"] = True
        elif emp in _NO:
            prof["has_employees"] = False
        note = " / ".join(x for x in (업종, 비고) if x)
        if note:
            prof["note"] = note
        out[name] = prof
    return out


def to_yaml(profiles: dict[str, dict], *, defaults: dict | None = None) -> str:
    """조사표 결과 → clients.yaml 텍스트."""
    import yaml

    doc = {
        "defaults": defaults or {"client_type": "corporate", "has_employees": True},
        "clients": profiles,
    }
    header = (
        "# config/clients.yaml — 수임처 속성 (kafa-clients import 로 생성)\n"
        "# 담당자만 아는 사실이라 조사표(엑셀)로 받아 변환한다.\n"
        "# 배경·근거: docs/domain_notes.md\n\n"
    )
    return header + yaml.safe_dump(doc, allow_unicode=True, sort_keys=False)
